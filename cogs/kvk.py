import io
import openpyxl
import discord
from discord import app_commands
from discord.ext import commands
from datetime import datetime, timezone

from config import COLOR_BOT, ALIANZA_FULL, ALIANZA_TAG, REINO
from checks import solo_en_canal
from db import database as db


class ConfirmResetView(discord.ui.View):
    def __init__(self, temporada_id: int, temporada_nombre: str):
        super().__init__(timeout=30)
        self.temporada_id     = temporada_id
        self.temporada_nombre = temporada_nombre

    @discord.ui.button(label='Sí, borrar todo', style=discord.ButtonStyle.danger, emoji='🗑️')
    async def confirmar(self, interaction: discord.Interaction, button: discord.ui.Button):
        await db.kvk_clear_import(self.temporada_id)
        await db.kvk_clear_stats(self.temporada_id)
        self.stop()
        await interaction.response.edit_message(
            embed=discord.Embed(
                title='✅ KvK reseteado',
                description=f'Todos los datos de **{self.temporada_nombre}** han sido borrados.\nPuedes usar `/kvk-importar` para cargar datos nuevos.',
                color=COLOR_BOT,
            ),
            view=None,
        )

    @discord.ui.button(label='Cancelar', style=discord.ButtonStyle.secondary, emoji='✖️')
    async def cancelar(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.stop()
        await interaction.response.edit_message(
            embed=discord.Embed(title='❌ Reset cancelado', color=COLOR_BOT),
            view=None,
        )


def parse_excel(data: bytes) -> list[dict]:
    wb  = openpyxl.load_workbook(io.BytesIO(data))

    # Buscar hoja K{id} (datos de jugadores individuales)
    hoja = next((wb[s] for s in wb.sheetnames if s.startswith('K') and s[1:].isdigit()), None)
    if not hoja:
        raise ValueError('No encontré la hoja de jugadores (K####) en el Excel.')

    # Fila 4 = cabeceras, fila 5+ = datos
    # Columnas 27-38 (1-based) = sección Deltas
    # En 0-indexed desde la tupla de valores: 26-37
    jugadores = []
    for row in hoja.iter_rows(min_row=5, values_only=True):
        gov_id    = row[26]
        gov_name  = row[27]
        muertes   = row[30]
        kill_pts  = row[31]
        kills_t4  = row[35]
        kills_t5  = row[36]
        dkp       = row[37]

        if not gov_id or not gov_name:
            continue
        if (kills_t4 or 0) < 0 and (kills_t5 or 0) < 0:
            continue  # Datos negativos = jugador inactivo/excluido

        jugadores.append({
            'governor_id':   str(gov_id),
            'governor_name': str(gov_name),
            'kills_t4':      max(0, kills_t4 or 0),
            'kills_t5':      max(0, kills_t5 or 0),
            'kill_points':   max(0, kill_pts  or 0),
            'muertes':       max(0, muertes   or 0),
            'dkp':           max(0, dkp       or 0),
        })

    return jugadores


def build_recap_embed(temporada, stats: list) -> discord.Embed:
    medallas = ['🥇', '🥈', '🥉']
    dkp_ranking = sorted(stats, key=lambda s: s['dkp'], reverse=True)

    total_t4  = sum(s['kills_t4']    for s in stats)
    total_t5  = sum(s['kills_t5']    for s in stats)
    total_kp  = sum(s['kill_points'] for s in stats)
    total_mts = sum(s['muertes']     for s in stats)

    embed = discord.Embed(
        title=f'🔥 Resumen de Temporada — {temporada["nombre"]}',
        description=(
            f'**{ALIANZA_FULL} · Reino {REINO}**\n'
            f'👥 **{len(stats)}** gobernadores participaron en este KvK'
        ),
        color=0xFF4444,
        timestamp=datetime.now(timezone.utc),
    )

    # Top 3 kills
    lineas_kills = []
    for i, s in enumerate(stats[:3]):
        total = s['kills_t4'] + s['kills_t5']
        lineas_kills.append(
            f'{medallas[i]} **{s["governor_name"]}**\n'
            f'`T4: {s["kills_t4"]:,}  ·  T5: {s["kills_t5"]:,}  ·  Total: {total:,}`'
        )
    embed.add_field(name='⚔️ Top Kills', value='\n'.join(lineas_kills) or '—', inline=False)

    # Top 3 DKP
    top_dkp = [s for s in dkp_ranking[:3] if s['dkp'] > 0]
    if top_dkp:
        lineas_dkp = [
            f'{medallas[i]} **{s["governor_name"]}** — `{s["dkp"]:,}` DKP'
            for i, s in enumerate(top_dkp)
        ]
        embed.add_field(name='💰 Top DKP', value='\n'.join(lineas_dkp), inline=False)

    # Estadísticas globales de la alianza
    embed.add_field(
        name='📊 La alianza en cifras',
        value=(
            f'⚔️ Kills T4 · **{total_t4:,}**\n'
            f'🏆 Kills T5 · **{total_t5:,}**\n'
            f'🎯 Kill Points · **{total_kp:,}**\n'
            f'💀 Muertes · **{total_mts:,}**'
        ),
        inline=True,
    )

    # MVP
    if stats:
        mvp = stats[0]
        mvp_total = mvp['kills_t4'] + mvp['kills_t5']
        embed.add_field(
            name='🏅 MVP del KvK',
            value=f'**{mvp["governor_name"]}**\n`{mvp_total:,}` kills totales',
            inline=True,
        )

    embed.set_footer(text=f'{ALIANZA_TAG} · Reino {REINO} · Temporada cerrada')
    return embed


class Kvk(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot

    @app_commands.command(name='kvk-inicio', description='[ADMIN] Inicia una nueva temporada KvK')
    @app_commands.describe(nombre='Nombre de la temporada (ej: KvK Mayo 2026)')
    @app_commands.checks.has_permissions(manage_guild=True)
    async def kvk_inicio(self, interaction: discord.Interaction, nombre: str):
        # Cerrar temporada anterior si existe (sin borrar sus datos históricos)
        await db.kvk_end_active(str(interaction.guild_id))
        temporada_id = await db.kvk_create(str(interaction.guild_id), nombre)
        embed = discord.Embed(
            title='⚔️ Nuevo KvK iniciado',
            description=f'**{nombre}** — ID `{temporada_id}`',
            color=0xFF4444,
        )
        embed.add_field(
            name='Próximo paso',
            value='Exporta el Excel de [heroscroll.com](https://heroscroll.com/rok/kvk-dashboard) al final del KvK y usa `/kvk-importar`',
            inline=False,
        )
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name='kvk-fin', description='[ADMIN] Cierra la temporada KvK activa y publica el resumen')
    @app_commands.checks.has_permissions(manage_guild=True)
    async def kvk_fin(self, interaction: discord.Interaction):
        temporada = await db.kvk_get_active(str(interaction.guild_id))
        if not temporada:
            await interaction.response.send_message('❌ No hay ninguna temporada activa.', ephemeral=True)
            return

        await interaction.response.defer()

        stats = await db.kvk_get_import_ranking(temporada['id'])
        await db.kvk_end_active(str(interaction.guild_id))

        await interaction.followup.send(f'✅ Temporada **{temporada["nombre"]}** cerrada.')

        if not stats:
            return

        # Buscar canal de recap: primero config guardada, luego kvk-anuncios por nombre
        canal = None
        canal_id = await db.get_config(str(interaction.guild_id), 'kvk_canal_recap')
        if canal_id:
            canal = interaction.guild.get_channel(int(canal_id))
        if not canal:
            canal = next(
                (c for c in interaction.guild.text_channels if 'kvk-anuncios' in c.name),
                None,
            )
        if canal:
            await canal.send(embed=build_recap_embed(temporada, stats))

    @app_commands.command(name='kvk-importar', description='[ADMIN] Importa estadísticas KvK desde el Excel de heroscroll.com')
    @app_commands.describe(archivo='Archivo .xlsx exportado de heroscroll.com/rok/kvk-dashboard')
    @app_commands.checks.has_permissions(manage_guild=True)
    async def kvk_importar(self, interaction: discord.Interaction, archivo: discord.Attachment):
        if not archivo.filename.endswith('.xlsx'):
            await interaction.response.send_message('❌ El archivo debe ser un `.xlsx`.', ephemeral=True)
            return

        temporada = await db.kvk_get_active(str(interaction.guild_id))
        if not temporada:
            await interaction.response.send_message(
                '❌ No hay temporada KvK activa. Usa `/kvk-inicio` primero.', ephemeral=True
            )
            return

        await interaction.response.defer()

        try:
            data      = await archivo.read()
            jugadores = parse_excel(data)
        except Exception as e:
            await interaction.followup.send(f'❌ Error leyendo el Excel: `{e}`', ephemeral=True)
            return

        if not jugadores:
            await interaction.followup.send('❌ No se encontraron jugadores en el archivo.', ephemeral=True)
            return

        # Insertar en BD
        rows = [
            (
                temporada['id'],
                j['governor_id'],
                j['governor_name'],
                j['kills_t4'],
                j['kills_t5'],
                j['kill_points'],
                j['muertes'],
                j['dkp'],
            )
            for j in jugadores
        ]
        await db.kvk_clear_import(temporada['id'])
        await db.kvk_insert_import(temporada['id'], rows)

        # Top 5 para el embed de resumen
        top5 = sorted(jugadores, key=lambda j: j['kills_t4'] + j['kills_t5'], reverse=True)[:5]
        medallas = ['🥇', '🥈', '🥉', '4️⃣', '5️⃣']

        embed = discord.Embed(
            title=f'✅ KvK importado — {temporada["nombre"]}',
            description=f'**{len(jugadores)}** jugadores cargados desde `{archivo.filename}`',
            color=0xFF4444,
        )
        lineas = []
        for i, j in enumerate(top5):
            total = j['kills_t4'] + j['kills_t5']
            lineas.append(
                f'{medallas[i]} **{j["governor_name"]}** — '
                f'T4: `{j["kills_t4"]:,}` · T5: `{j["kills_t5"]:,}` · '
                f'Total: `{total:,}` · 💀 `{j["muertes"]:,}`'
            )
        embed.add_field(name='🏆 Top 5', value='\n'.join(lineas), inline=False)
        embed.set_footer(text='Usa /kvk-ranking para ver el ranking completo')
        await interaction.followup.send(embed=embed)

    @app_commands.command(name='kvk-ranking', description='[ADMIN] Muestra el ranking del KvK actual')
    @solo_en_canal('kvk')
    @app_commands.checks.has_permissions(manage_guild=True)
    async def kvk_ranking(self, interaction: discord.Interaction):
        temporada = await db.kvk_get_active(str(interaction.guild_id))
        if not temporada:
            await interaction.response.send_message('❌ No hay ninguna temporada KvK activa.', ephemeral=True)
            return

        # Usar datos importados si existen, si no los manuales
        stats = await db.kvk_get_import_ranking(temporada['id'])
        fuente = '📊 heroscroll.com'

        if not stats:
            stats  = await db.kvk_get_ranking(temporada['id'])
            fuente = '✏️ Registros manuales'

        if not stats:
            await interaction.response.send_message(
                '❌ Sin datos. Usa `/kvk-importar` con el Excel de heroscroll.com.',
                ephemeral=True,
            )
            return

        medallas = ['🥇', '🥈', '🥉']
        lineas = []
        for i, s in enumerate(stats):
            t4    = s['kills_t4']
            t5    = s['kills_t5']
            total = t4 + t5
            name  = s.get('governor_name') or s.get('username', '?')
            pos   = medallas[i] if i < 3 else f'**{i+1}.**'
            lineas.append(
                f'{pos} **{name}** — '
                f'T4: `{t4:,}` · T5: `{t5:,}` · '
                f'Total: `{total:,}` · 💀 `{s["muertes"]:,}`'
            )

        # Dividir en páginas de 20 si hay muchos jugadores
        paginas = [lineas[i:i+20] for i in range(0, len(lineas), 20)]
        embed = discord.Embed(
            title=f'🏆 Ranking KvK — {temporada["nombre"]}',
            description='\n'.join(paginas[0]),
            color=0xFF4444,
        )
        embed.set_footer(text=f'Fuente: {fuente} · {len(stats)} jugadores')
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name='kvk-buscar', description='Busca las stats de un gobernador por nombre')
    @solo_en_canal('kvk')
    @app_commands.describe(nombre='Nombre del gobernador (parcial también funciona)')
    async def kvk_buscar(self, interaction: discord.Interaction, nombre: str):
        temporada = await db.kvk_get_active(str(interaction.guild_id))
        if not temporada:
            await interaction.response.send_message('❌ No hay temporada activa.', ephemeral=True)
            return

        stats = await db.kvk_get_import_ranking(temporada['id'])
        if not stats:
            await interaction.response.send_message('❌ No hay datos importados.', ephemeral=True)
            return

        encontrados = [
            s for s in stats
            if nombre.lower() in s['governor_name'].lower()
        ]

        if not encontrados:
            await interaction.response.send_message(
                f'❌ No encontré ningún gobernador con "{nombre}".', ephemeral=True
            )
            return

        s = encontrados[0]
        total = s['kills_t4'] + s['kills_t5']
        ranking_pos = next((i+1 for i, r in enumerate(stats) if r['governor_id'] == s['governor_id']), '?')

        embed = discord.Embed(
            title=f'⚔️ Stats KvK — {s["governor_name"]}',
            color=0xFF4444,
        )
        embed.add_field(name='🏅 Posición',   value=f'#{ranking_pos}', inline=True)
        embed.add_field(name='T4 kills',      value=f'{s["kills_t4"]:,}', inline=True)
        embed.add_field(name='T5 kills',      value=f'{s["kills_t5"]:,}', inline=True)
        embed.add_field(name='Total kills',   value=f'{total:,}', inline=True)
        embed.add_field(name='💀 Muertes',    value=f'{s["muertes"]:,}', inline=True)
        embed.add_field(name='Kill Points',   value=f'{s["kill_points"]:,}', inline=True)
        embed.add_field(name='DKP',           value=f'{s["dkp"]:,}', inline=True)
        embed.set_footer(text=f'Temporada: {temporada["nombre"]}')
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name='kvk-reset', description='[ADMIN] Borra todos los datos del KvK actual y permite empezar de cero')
    @app_commands.checks.has_permissions(manage_guild=True)
    async def kvk_reset(self, interaction: discord.Interaction):
        temporada = await db.kvk_get_active(str(interaction.guild_id))
        if not temporada:
            await interaction.response.send_message('❌ No hay ninguna temporada activa.', ephemeral=True)
            return

        # Confirmación con botones
        embed = discord.Embed(
            title='⚠️ ¿Confirmar reset del KvK?',
            description=(
                f'Vas a borrar **todos los datos** de la temporada:\n'
                f'**{temporada["nombre"]}**\n\n'
                f'Esto eliminará:\n'
                f'• Todos los datos importados del Excel\n'
                f'• Todos los registros manuales\n\n'
                f'La temporada seguirá activa para que puedas volver a importar.'
            ),
            color=0xFF4444,
        )

        view = ConfirmResetView(temporada['id'], temporada['nombre'])
        await interaction.response.send_message(embed=embed, view=view, ephemeral=True)

    @app_commands.command(name='kvk-dkp', description='Ranking de DKP de la temporada actual')
    @solo_en_canal('kvk')
    async def kvk_dkp(self, interaction: discord.Interaction):
        temporada = await db.kvk_get_active(str(interaction.guild_id))
        if not temporada:
            await interaction.response.send_message('❌ No hay temporada activa.', ephemeral=True)
            return

        stats = await db.kvk_get_import_ranking(temporada['id'])
        if not stats:
            await interaction.response.send_message('❌ No hay datos importados. Usa `/kvk-importar` primero.', ephemeral=True)
            return

        ranking = sorted((s for s in stats if s['dkp'] > 0), key=lambda s: s['dkp'], reverse=True)
        if not ranking:
            await interaction.response.send_message('❌ No hay datos de DKP en esta temporada.', ephemeral=True)
            return

        medallas = ['🥇', '🥈', '🥉']
        lineas = []
        for i, s in enumerate(ranking):
            pos   = medallas[i] if i < 3 else f'**{i+1}.**'
            total = s['kills_t4'] + s['kills_t5']
            lineas.append(
                f'{pos} **{s["governor_name"]}** — '
                f'`{s["dkp"]:,}` DKP · '
                f'Kills: `{total:,}` · 💀 `{s["muertes"]:,}`'
            )

        embed = discord.Embed(
            title=f'💰 Ranking DKP — {temporada["nombre"]}',
            description='\n'.join(lineas[:20]),
            color=0xFFD700,
        )
        embed.set_footer(text=f'{len(ranking)} gobernadores con DKP · {ALIANZA_TAG} · Reino {REINO}')
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name='kvk-canal-recap', description='[ADMIN] Configura el canal donde se publica el resumen al cerrar el KvK')
    @app_commands.describe(canal='Canal donde se publicará el resumen de fin de temporada')
    @app_commands.checks.has_permissions(manage_guild=True)
    async def kvk_canal_recap(self, interaction: discord.Interaction, canal: discord.TextChannel):
        await db.set_config(str(interaction.guild_id), 'kvk_canal_recap', str(canal.id))
        await interaction.response.send_message(
            f'✅ El resumen de fin de temporada se publicará en {canal.mention}.',
            ephemeral=True,
        )

    @kvk_inicio.error
    @kvk_fin.error
    @kvk_importar.error
    @kvk_reset.error
    @kvk_ranking.error
    @kvk_canal_recap.error
    async def admin_error(self, interaction: discord.Interaction, error: app_commands.AppCommandError):
        if isinstance(error, app_commands.MissingPermissions):
            await interaction.response.send_message('❌ No tienes permisos para este comando.', ephemeral=True)


async def setup(bot: commands.Bot):
    await bot.add_cog(Kvk(bot))
